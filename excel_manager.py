from io import BytesIO
from openpyxl import Workbook, load_workbook

def create_excel_with_headers(headers: list) -> Workbook:
    """
    Create a new Excel workbook with headers in the first row.
    
    Args:
        headers: List of column header names
        
    Returns:
        A new openpyxl Workbook with headers
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "VENTAS DHELOS"
    
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    return wb


def update_excel(flat_data: dict, file_id: str, drive_service: Resource) -> dict:
    """
    Main orchestrator: downloads Excel, appends row, uploads back.
    
    If the file doesn't exist or is empty, creates a new Excel with headers.
    
    Args:
        flat_data: Flattened dictionary to append as a row
        file_id: Google Drive file ID
        
    Returns:
        Dict with status and message
    """
    
    try:
        # Try to download existing file
        buffer = download_excel(drive_service, file_id)
        wb = load_workbook(buffer)
        ws = wb.active
        
        # Check if file has headers, if not, add them
        if ws.max_row == 0 or ws.cell(row=1, column=1).value is None:
            headers = list(flat_data.keys())
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)
            next_row = 2
        else:
            next_row = ws.max_row + 1
            # Get headers from first row
            headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        
    except Exception as e:
        # File doesn't exist or can't be read - create new
        headers = list(flat_data.keys())
        wb = create_excel_with_headers(headers)
        ws = wb.active
        next_row = 2
    
    # Append data row based on headers order
    for col_idx, header in enumerate(headers, start=1):
        value = flat_data.get(header, "")
        ws.cell(row=next_row, column=col_idx, value=value)
    
    # Upload updated workbook
    upload_excel(drive_service, file_id, wb)
    
    return {
        "status": "success",
        "message": f"Data appended to row {next_row}",
        "row": next_row
    }
